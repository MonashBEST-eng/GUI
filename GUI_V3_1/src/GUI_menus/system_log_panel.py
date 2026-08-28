import tkinter as tk
from tkinter import filedialog, messagebox
import os
import time
import re
from datetime import datetime, timedelta

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


BG_MAIN  = "#0f172a"
BG_CARD  = "#1f2937"
BG_PANEL = "#111827"
TEXT     = "#e5e7eb"

LOG_PATH = os.path.join("outputs", "system_log.txt")


def open_system_log_panel(root):
    panel = tk.Toplevel(root)
    panel.title("System Log")
    panel.geometry("900x650")
    panel.configure(bg=BG_MAIN)

    refresh_job = None

    # =========================
    # HEADER
    # =========================
    header = tk.Frame(panel, bg=BG_PANEL)
    header.pack(fill="x")

    tk.Label(
        header,
        text="SYSTEM LOG",
        fg="white",
        bg=BG_PANEL,
        font=("Segoe UI", 16, "bold")
    ).pack(side="left", padx=16, pady=12)

    status_var = tk.StringVar(value="")

    tk.Label(
        header,
        textvariable=status_var,
        fg="gray",
        bg=BG_PANEL,
        font=("Segoe UI", 9)
    ).pack(side="right", padx=16)

    # =========================
    # MAIN CARD
    # =========================
    card = tk.Frame(panel, bg=BG_CARD)
    card.pack(fill="both", expand=True, padx=16, pady=16)

    tk.Label(
        card,
        text="📄 outputs/system_log.txt",
        fg="cyan",
        bg=BG_CARD,
        font=("Segoe UI", 10, "bold")
    ).pack(anchor="w", padx=12, pady=(10, 4))

    # =========================
    # TEXT BOX + SCROLLBAR
    # =========================
    text_frame = tk.Frame(card, bg=BG_CARD)
    text_frame.pack(fill="both", expand=True, padx=12, pady=(0, 12))

    scrollbar = tk.Scrollbar(text_frame)
    scrollbar.pack(side="right", fill="y")

    log_box = tk.Text(
        text_frame,
        bg=BG_MAIN,
        fg=TEXT,
        font=("Consolas", 9),
        wrap="none",
        yscrollcommand=scrollbar.set
    )
    log_box.pack(side="left", fill="both", expand=True)

    scrollbar.config(command=log_box.yview)

    # Colour tags
    log_box.tag_config("error", foreground="#ef4444")
    log_box.tag_config("warn", foreground="#fb923c")
    log_box.tag_config("info", foreground="#e5e7eb")
    log_box.tag_config("cmd", foreground="#c084fc")
    log_box.tag_config("rx", foreground="#38bdf8")
    log_box.tag_config("tx", foreground="#a3e635")
    log_box.tag_config("conn", foreground="#fb923c")
    log_box.tag_config("btn", foreground="#c084fc")
    log_box.tag_config("alarm", foreground="#ef4444")

    # =========================
    # LOG PARSING
    # =========================
    def get_line_tag(line):
        lower = line.lower()

        if "error" in lower or "alarm" in lower or "[alarm]" in lower:
            return "alarm"
        elif "warn" in lower or "[warn]" in lower:
            return "warn"
        elif "[rx]" in lower:
            return "rx"
        elif "[tx]" in lower:
            return "tx"
        elif "[conn]" in lower or "[mqtt]" in lower:
            return "conn"
        elif "[btn]" in lower or "[cmd]" in lower or "button" in lower:
            return "btn"
        else:
            return "info"

    def parse_log_line(line):
        """
        Supports lines like:
        [RX] [10:35:14] tbm/topic | payload
        [CONN] [10:35:14] [MQTT] Disconnected
        [10:35:14] normal message
        """

        clean = line.strip()

        category = ""
        timestamp = ""
        message = clean

        # Match optional category then timestamp
        match = re.match(
            r"^\[(?P<first>[^\]]+)\]\s*(?:\[(?P<second>\d{2}:\d{2}:\d{2})\])?\s*(?P<rest>.*)$",
            clean
        )

        if match:
            first = match.group("first")
            second = match.group("second")
            rest = match.group("rest")

            if re.match(r"^\d{2}:\d{2}:\d{2}$", first):
                timestamp = first
                category = ""
                message = rest
            elif second:
                category = first
                timestamp = second
                message = rest
            else:
                category = first
                message = rest

        return {
            "timestamp": timestamp,
            "category": category,
            "message": message,
            "raw": clean
        }

    def read_log_lines():
        if not os.path.exists(LOG_PATH):
            return []

        with open(LOG_PATH, "r", encoding="utf-8", errors="ignore") as file:
            return file.readlines()

    def timestamp_to_datetime_today(timestamp):
        """
        Converts HH:MM:SS to a datetime using today's date.
        If the log crossed midnight, this can be imperfect because the log file
        only stores time, not date.
        """
        if not timestamp:
            return None

        try:
            parsed_time = datetime.strptime(timestamp, "%H:%M:%S").time()
            today = datetime.now().date()
            return datetime.combine(today, parsed_time)
        except ValueError:
            return None

    def is_line_in_past_minutes(line, minutes):
        """
        Returns True if a log line occurred within the past X minutes.
        Uses today's date because the log only stores HH:MM:SS.
        """

        now = datetime.now()
        cutoff = now - timedelta(minutes=minutes)

        parsed = parse_log_line(line)
        log_dt = timestamp_to_datetime_today(parsed["timestamp"])

        if log_dt is None:
            return False

        # Handle simple midnight rollover case:
        # If a log time appears in the future compared to now, assume it was yesterday.
        if log_dt > now:
            log_dt = log_dt - timedelta(days=1)

        return log_dt >= cutoff

    def get_lines_from_past_minutes(log_lines, minutes):
        """
        Returns log lines with timestamps within the past X minutes.
        Uses today's date because the log only stores HH:MM:SS.
        """

        selected_lines = []

        for line in log_lines:
            if is_line_in_past_minutes(line, minutes):
                selected_lines.append(line)

        return selected_lines

    # =========================
    # LOG LOADING
    # =========================
    def load_log():
        log_box.config(state="normal")
        log_box.delete("1.0", tk.END)

        if not os.path.exists(LOG_PATH):
            log_box.insert(tk.END, "No system log file found yet.\n", "warn")
            log_box.insert(tk.END, f"Expected path: {LOG_PATH}\n", "info")
        else:
            with open(LOG_PATH, "r", encoding="utf-8", errors="ignore") as file:
                for line in file:
                    tag = get_line_tag(line)
                    log_box.insert(tk.END, line, tag)

        log_box.see(tk.END)
        log_box.config(state="disabled")
        status_var.set(f"Last refresh {time.strftime('%H:%M:%S')}")

    # =========================
    # CLEAR LOG FILE
    # =========================
    ask_before_clearing_enabled = tk.BooleanVar(value=False)

    def clear_log_file():
        if ask_before_clearing_enabled.get():
            confirm = messagebox.askyesno(
                "Clear System Log",
                "This will permanently clear outputs/system_log.txt.\n\nAre you sure?"
            )

            if not confirm:
                return

        log_dir = os.path.dirname(LOG_PATH)
        if log_dir:
            os.makedirs(log_dir, exist_ok=True)

        with open(LOG_PATH, "w", encoding="utf-8") as file:
            file.write("")

        log_box.config(state="normal")
        log_box.delete("1.0", tk.END)
        log_box.config(state="disabled")

        status_var.set(f"Log cleared at {time.strftime('%H:%M:%S')}")

    # =========================
    # SAVE TO EXCEL
    # =========================
    def choose_export_range(log_lines):
        """
        Opens a popup where the user can choose:
        - manual start/end log range
        - or export logs from the past X minutes

        Defaults:
        - start = first log line
        - end = last log line
        - past X minutes disabled
        """

        parsed_lines = [parse_log_line(line) for line in log_lines]

        chooser = tk.Toplevel(panel)
        chooser.title("Choose Export Range")
        chooser.geometry("900x650")
        chooser.configure(bg=BG_MAIN)
        chooser.grab_set()

        selected_range = {
            "mode": None,
            "start": None,
            "end": None,
            "minutes": None
        }

        use_past_minutes = tk.BooleanVar(value=False)
        past_minutes_value = tk.StringVar(value="10")
        highlight_count_var = tk.StringVar(value="")

        tk.Label(
            chooser,
            text="Choose log range to export:",
            fg="white",
            bg=BG_MAIN,
            font=("Segoe UI", 11, "bold")
        ).pack(anchor="w", padx=14, pady=(14, 6))

        tk.Label(
            chooser,
            text="All log entries are selected by default. Change the start/end entries, or tick Past X minutes.",
            fg="gray",
            bg=BG_MAIN,
            font=("Segoe UI", 9)
        ).pack(anchor="w", padx=14, pady=(0, 10))

        # =========================
        # PAST X MINUTES OPTION
        # =========================
        past_frame = tk.Frame(chooser, bg=BG_CARD)
        past_frame.pack(fill="x", padx=14, pady=(0, 8))

        def reset_list_highlighting():
            for i in range(len(log_lines)):
                start_list.itemconfig(i, bg=BG_PANEL, fg=TEXT)
                end_list.itemconfig(i, bg=BG_PANEL, fg=TEXT)

            highlight_count_var.set("")

        def highlight_past_minutes_lines():
            """
            Highlights rows in the existing Start Entry and End Entry boxes
            which fall inside the past X minutes.
            """

            reset_list_highlighting()

            if not use_past_minutes.get():
                return

            try:
                minutes = float(past_minutes_value.get())
            except ValueError:
                highlight_count_var.set("Enter a valid number")
                return

            if minutes <= 0:
                highlight_count_var.set("Minutes must be > 0")
                return

            matching_indices = []

            for i, line in enumerate(log_lines):
                if is_line_in_past_minutes(line, minutes):
                    matching_indices.append(i)
                    start_list.itemconfig(i, bg="#14532d", fg="#bbf7d0")
                    end_list.itemconfig(i, bg="#14532d", fg="#bbf7d0")

            if matching_indices:
                first_idx = matching_indices[0]
                last_idx = matching_indices[-1]

                start_list.selection_clear(0, tk.END)
                end_list.selection_clear(0, tk.END)

                start_list.selection_set(first_idx)
                start_list.activate(first_idx)
                start_list.see(first_idx)

                end_list.selection_set(last_idx)
                end_list.activate(last_idx)
                end_list.see(last_idx)

                highlight_count_var.set(f"{len(matching_indices)} matching log entries highlighted")
            else:
                highlight_count_var.set("No matching log entries found")

        def update_past_minutes_state():
            state = "normal" if use_past_minutes.get() else "disabled"
            past_minutes_entry.config(state=state)
            highlight_past_minutes_lines()

        tk.Checkbutton(
            past_frame,
            text="Export past",
            variable=use_past_minutes,
            bg=BG_CARD,
            fg="white",
            selectcolor=BG_PANEL,
            activebackground=BG_CARD,
            activeforeground="white",
            font=("Segoe UI", 9),
            command=update_past_minutes_state
        ).pack(side="left", padx=(10, 4), pady=8)

        past_minutes_entry = tk.Entry(
            past_frame,
            textvariable=past_minutes_value,
            bg=BG_PANEL,
            fg=TEXT,
            insertbackground=TEXT,
            width=8,
            font=("Segoe UI", 9),
            justify="center"
        )
        past_minutes_entry.pack(side="left", padx=4, pady=8)

        tk.Label(
            past_frame,
            text="minutes",
            fg="white",
            bg=BG_CARD,
            font=("Segoe UI", 9)
        ).pack(side="left", padx=(4, 10), pady=8)

        tk.Label(
            past_frame,
            text="When ticked, highlighted entries below will be exported.",
            fg="gray",
            bg=BG_CARD,
            font=("Segoe UI", 8)
        ).pack(side="left", padx=8, pady=8)

        tk.Label(
            past_frame,
            textvariable=highlight_count_var,
            fg="#bbf7d0",
            bg=BG_CARD,
            font=("Segoe UI", 8, "bold")
        ).pack(side="right", padx=10, pady=8)

        # =========================
        # RANGE SELECTION AREA
        # =========================
        range_frame = tk.Frame(chooser, bg=BG_MAIN)
        range_frame.pack(fill="both", expand=True, padx=14, pady=8)

        range_frame.columnconfigure(0, weight=1)
        range_frame.columnconfigure(1, weight=1)
        range_frame.rowconfigure(0, weight=1)

        # ---------- START LIST ----------
        start_frame = tk.Frame(range_frame, bg=BG_CARD)
        start_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 6))

        tk.Label(
            start_frame,
            text="Start Entry",
            fg="white",
            bg=BG_CARD,
            font=("Segoe UI", 10, "bold")
        ).pack(anchor="w", padx=10, pady=(8, 4))

        start_list_frame = tk.Frame(start_frame, bg=BG_CARD)
        start_list_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        start_scroll = tk.Scrollbar(start_list_frame)
        start_scroll.pack(side="right", fill="y")

        start_list = tk.Listbox(
            start_list_frame,
            bg=BG_PANEL,
            fg=TEXT,
            font=("Consolas", 8),
            selectbackground="#2563eb",
            selectforeground="white",
            yscrollcommand=start_scroll.set,
            exportselection=False
        )
        start_list.pack(side="left", fill="both", expand=True)
        start_scroll.config(command=start_list.yview)

        # ---------- END LIST ----------
        end_frame = tk.Frame(range_frame, bg=BG_CARD)
        end_frame.grid(row=0, column=1, sticky="nsew", padx=(6, 0))

        tk.Label(
            end_frame,
            text="End Entry",
            fg="white",
            bg=BG_CARD,
            font=("Segoe UI", 10, "bold")
        ).pack(anchor="w", padx=10, pady=(8, 4))

        end_list_frame = tk.Frame(end_frame, bg=BG_CARD)
        end_list_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        end_scroll = tk.Scrollbar(end_list_frame)
        end_scroll.pack(side="right", fill="y")

        end_list = tk.Listbox(
            end_list_frame,
            bg=BG_PANEL,
            fg=TEXT,
            font=("Consolas", 8),
            selectbackground="#2563eb",
            selectforeground="white",
            yscrollcommand=end_scroll.set,
            exportselection=False
        )
        end_list.pack(side="left", fill="both", expand=True)
        end_scroll.config(command=end_list.yview)

        # ---------- POPULATE BOTH LISTS ----------
        for i, item in enumerate(parsed_lines):
            timestamp = item["timestamp"] if item["timestamp"] else "--:--:--"
            category = item["category"] if item["category"] else "LOG"
            preview = item["message"]

            if len(preview) > 65:
                preview = preview[:65] + "..."

            line_text = f"{i + 1:04d} | {timestamp} | {category:<6} | {preview}"

            start_list.insert(tk.END, line_text)
            end_list.insert(tk.END, line_text)

        # Default: export all logs
        if log_lines:
            start_list.selection_set(0)
            start_list.activate(0)
            start_list.see(0)

            end_index = len(log_lines) - 1
            end_list.selection_set(end_index)
            end_list.activate(end_index)
            end_list.see(end_index)

        # Update highlighting whenever the minutes entry changes
        past_minutes_value.trace_add("write", lambda *_: highlight_past_minutes_lines())

        # Start disabled
        past_minutes_entry.config(state="disabled")

        # =========================
        # BUTTONS
        # =========================
        button_row = tk.Frame(chooser, bg=BG_MAIN)
        button_row.pack(fill="x", padx=14, pady=(4, 14))

        def confirm_selection():
            if use_past_minutes.get():
                try:
                    minutes = float(past_minutes_value.get())
                except ValueError:
                    messagebox.showwarning(
                        "Invalid Minutes",
                        "Please enter a valid number of minutes."
                    )
                    return

                if minutes <= 0:
                    messagebox.showwarning(
                        "Invalid Minutes",
                        "Minutes must be greater than 0."
                    )
                    return

                selected_range["mode"] = "past_minutes"
                selected_range["minutes"] = minutes
                chooser.destroy()
                return

            start_selection = start_list.curselection()
            end_selection = end_list.curselection()

            if not start_selection or not end_selection:
                messagebox.showwarning(
                    "No Selection",
                    "Please select both a start entry and an end entry."
                )
                return

            start_index = start_selection[0]
            end_index = end_selection[0]

            if start_index > end_index:
                messagebox.showwarning(
                    "Invalid Range",
                    "The start entry must be before or equal to the end entry."
                )
                return

            selected_range["mode"] = "manual_range"
            selected_range["start"] = start_index
            selected_range["end"] = end_index
            chooser.destroy()

        def cancel_selection():
            selected_range["mode"] = None
            selected_range["start"] = None
            selected_range["end"] = None
            selected_range["minutes"] = None
            chooser.destroy()

        tk.Button(
            button_row,
            text="Export",
            bg=BG_PANEL,
            fg="white",
            relief="flat",
            padx=16,
            pady=8,
            command=confirm_selection
        ).pack(side="left")

        tk.Button(
            button_row,
            text="Cancel",
            bg=BG_PANEL,
            fg="white",
            relief="flat",
            padx=16,
            pady=8,
            command=cancel_selection
        ).pack(side="left", padx=8)

        chooser.wait_window()

        if selected_range["mode"] is None:
            return None

        return selected_range

    def save_to_excel():
        if Workbook is None:
            messagebox.showerror(
                "Missing Package",
                "openpyxl is not installed.\n\nInstall it from a terminal with:\n\npip install openpyxl"
            )
            return

        log_lines = read_log_lines()

        if not log_lines:
            messagebox.showwarning("No Log Data", "There is no system log data to save.")
            return

        export_selection = choose_export_range(log_lines)

        if export_selection is None:
            return

        if export_selection["mode"] == "past_minutes":
            selected_lines = get_lines_from_past_minutes(
                log_lines,
                export_selection["minutes"]
            )

            if not selected_lines:
                messagebox.showwarning(
                    "No Matching Logs",
                    f"No log entries were found in the past {export_selection['minutes']} minutes."
                )
                return

        else:
            start_index = export_selection["start"]
            end_index = export_selection["end"]
            selected_lines = log_lines[start_index:end_index + 1]

        save_path = filedialog.asksaveasfilename(
            parent=panel,
            title="Save System Log Excel File",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile=f"system_log_export_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if not save_path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "System Log"

        headers = ["Timestamp", "Category", "Message", "Raw Log Line"]
        ws.append(headers)

        # Header styling
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F2937")
            cell.alignment = Alignment(horizontal="center")

        for line in selected_lines:
            parsed = parse_log_line(line)

            ws.append([
                parsed["timestamp"],
                parsed["category"],
                parsed["message"],
                parsed["raw"]
            ])

        # Auto column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                value = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(value))

            # Message/raw columns can be wider
            if col_letter in ["C", "D"]:
                ws.column_dimensions[col_letter].width = min(max_length + 2, 80)
            else:
                ws.column_dimensions[col_letter].width = min(max_length + 2, 25)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Wrap message/raw columns
        for row in ws.iter_rows(min_row=2):
            row[2].alignment = Alignment(wrap_text=True, vertical="top")
            row[3].alignment = Alignment(wrap_text=True, vertical="top")

        try:
            wb.save(save_path)
            messagebox.showinfo(
                "Save Complete",
                f"System log saved successfully.\n\nSaved {len(selected_lines)} log entries to:\n{save_path}"
            )
        except Exception as e:
            messagebox.showerror("Save Failed", f"Could not save Excel file:\n\n{e}")

    # =========================
    # AUTO REFRESH
    # =========================
    auto_refresh_enabled = tk.BooleanVar(value=True)

    def auto_refresh():
        nonlocal refresh_job

        if not panel.winfo_exists():
            return

        if auto_refresh_enabled.get():
            load_log()
            refresh_job = panel.after(1000, auto_refresh)
        else:
            refresh_job = None

    def toggle_auto_refresh():
        nonlocal refresh_job

        if auto_refresh_enabled.get():
            if refresh_job is None:
                auto_refresh()
        else:
            if refresh_job is not None:
                panel.after_cancel(refresh_job)
                refresh_job = None

            status_var.set("Auto refresh paused")

    def on_close():
        nonlocal refresh_job

        if refresh_job is not None:
            panel.after_cancel(refresh_job)
            refresh_job = None

        panel.destroy()

    panel.protocol("WM_DELETE_WINDOW", on_close)

    # =========================
    # BUTTON ROW
    # =========================
    button_row = tk.Frame(panel, bg=BG_MAIN)
    button_row.pack(fill="x", padx=16, pady=(0, 16))

    tk.Button(
        button_row,
        text="Refresh",
        bg=BG_PANEL,
        fg="white",
        relief="flat",
        padx=16,
        pady=8,
        command=load_log
    ).pack(side="left")

    tk.Checkbutton(
        button_row,
        text="Auto refresh",
        variable=auto_refresh_enabled,
        bg=BG_MAIN,
        fg="white",
        selectcolor=BG_PANEL,
        activebackground=BG_MAIN,
        activeforeground="white",
        font=("Segoe UI", 9),
        command=toggle_auto_refresh
    ).pack(side="left", padx=10)

    tk.Button(
        button_row,
        text="Save to Excel",
        bg=BG_PANEL,
        fg="white",
        relief="flat",
        padx=16,
        pady=8,
        command=save_to_excel
    ).pack(side="left", padx=8)

    tk.Button(
        button_row,
        text="Clear Log",
        bg=BG_PANEL,
        fg="white",
        relief="flat",
        padx=16,
        pady=8,
        command=clear_log_file
    ).pack(side="left", padx=8)

    tk.Checkbutton(
        button_row,
        text="Ask before clearing log",
        variable=ask_before_clearing_enabled,
        bg=BG_MAIN,
        fg="white",
        selectcolor=BG_PANEL,
        activebackground=BG_MAIN,
        activeforeground="white",
        font=("Segoe UI", 9)
    ).pack(side="right", padx=8)

    auto_refresh()